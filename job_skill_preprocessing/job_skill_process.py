#for one job description, if this word exits, set it to 1, not set it to 0
# import pandas as pd
# import numpy as np


# data = pd.read_excel('user_profile.xlsx')
# train_data = np.array(data)
# excel_list = train_data.tolist()
# i = 1

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from nltk.tokenize import word_tokenize
import string
import re
import pandas as pd


# get all skills from user data
wb = openpyxl.load_workbook('user_profile.xlsx')

sheet = wb.active
max_column = sheet.max_column       
column = get_column_letter(max_column)

row = sheet['H2':'%s2' % column]   
 
skill_list = []

for row_cells in row:
    for cell in row_cells:
        skill_list.append(cell.value)

# print(skill_list)

# get all skills from company data
jobs = openpyxl.load_workbook('dice_com-job_us_sample.xlsx')

sheet = jobs.active
max_row = sheet.max_row

job_id = sheet['E1':'E%s' % max_row]


id_list = []

for cells in job_id:
    for cell in cells:
        id_list.append(cell.value)

job_des = sheet['D1':'D%s' % max_row]
des_list = []

for cells in job_des:
    for cell in cells:
        if cell.value != None:
            des_list.append(cell.value)
        else:
            des_list.append("")

job_skill = sheet['J1':'J%s' % max_row]

job_skill_list = []
for cells in job_skill:
    for cell in cells:
        if cell.value != None:
            job_skill_list.append(cell.value)
        else:
            job_skill_list.append("")

des_list = [s1 + ', ' + s2 for s1, s2 in zip(job_skill_list, des_list)]


# one-hot
res = {}

for skill in skill_list:
    res[skill] = []

i = 0


for des in des_list:
    i += 1
    word_tokens = word_tokenize(
        re.sub('[%s]' % re.escape(string.punctuation), ' ', des))
    temp_dict = {}
    for skill in skill_list:
        if skill in word_tokens:
            temp_dict[skill] = 1
        else:
            temp_dict[skill] = 0
    
    for skill in skill_list:
        if len(res[skill]) == i:
            continue
        res[skill].append(temp_dict[skill])
        
res['job id'] = id_list
df = pd.DataFrame(res)
# print(df)
df.to_csv("job_skill.csv", sep = ',')



        


    
    

